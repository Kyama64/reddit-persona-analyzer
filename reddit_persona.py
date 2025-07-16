import praw
import os
import re
import sys
import json
import nltk
import logging
import openpyxl
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from collections import Counter
from datetime import datetime, timezone
from dotenv import load_dotenv
import time

# Completely suppress NLTK download messages and warnings
import warnings
warnings.filterwarnings('ignore')

# Redirect NLTK download output to devnull
import contextlib
import io

with contextlib.redirect_stdout(io.StringIO()):
    with contextlib.redirect_stderr(io.StringIO()):
        try:
            nltk.data.find('tokenizers/punkt')
        except LookupError:
            nltk.download('punkt', quiet=True)

        try:
            nltk.data.find('corpora/stopwords')
        except LookupError:
            nltk.download('stopwords', quiet=True)

        try:
            nltk.data.find('sentiment/vader_lexicon')
        except LookupError:
            nltk.download('vader_lexicon', quiet=True)

# Set console encoding to UTF-8
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Load environment variables from .env file
load_dotenv()

# Download required NLTK data
try:
    nltk.data.find('tokenizers/punkt')
    nltk.data.find('corpora/stopwords')
    nltk.data.find('sentiment/vader_lexicon')
except LookupError:
    print("Downloading required NLTK data...")
    nltk.download('punkt', quiet=True)
    nltk.download('stopwords', quiet=True)
    nltk.download('vader_lexicon', quiet=True)

# Google Sheets API setup
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    import google.auth
    GOOGLE_SHEETS_AVAILABLE = True
except ImportError:
    GOOGLE_SHEETS_AVAILABLE = False

class RedditPersonaAnalyzer:
    def __init__(self, client_id, client_secret, user_agent, google_creds_path=None):
        """Initialize the Reddit client and Google Sheets if available"""
        # Initialize Reddit client
        self.reddit = praw.Reddit(
            client_id=client_id,
            client_secret=client_secret,
            user_agent=user_agent
        )
        
        # Initialize NLTK components
        self.sia = SentimentIntensityAnalyzer()
        self.stop_words = set(stopwords.words('english'))
        
        # Initialize Google Sheets if available
        self.google_sheets_service = None
        print(f"\nInitializing Google Sheets...")
        print(f"GOOGLE_SHEETS_AVAILABLE: {GOOGLE_SHEETS_AVAILABLE}")
        print(f"google_creds_path: {google_creds_path}")
        print(f"File exists: {os.path.exists(google_creds_path) if google_creds_path else 'N/A'}")
        
        if GOOGLE_SHEETS_AVAILABLE and google_creds_path and os.path.exists(google_creds_path):
            try:
                print("Attempting to create credentials...")
                creds = service_account.Credentials.from_service_account_file(
                    google_creds_path,
                    scopes=['https://www.googleapis.com/auth/spreadsheets']
                )
                print("Credentials created, building service...")
                self.google_sheets_service = build('sheets', 'v4', credentials=creds)
                print("✅ Google Sheets service initialized successfully!")
            except Exception as e:
                print(f"❌ Error initializing Google Sheets: {str(e)}")
                print(f"Error type: {type(e).__name__}")
                if hasattr(e, 'message'):
                    print(f"Error message: {e.message}")
                if hasattr(e, 'details'):
                    print(f"Error details: {e.details}")
    
    def export_to_google_sheets(self, username, persona_data, spreadsheet_id=None):
        """Export persona data to Google Sheets"""
        print("\nAttempting to export to Google Sheets...")
        if not self.google_sheets_service:
            print("❌ Google Sheets export is not configured. Please check your credentials.")
            print(f"google_sheets_service available: {hasattr(self, 'google_sheets_service')}")
            return None
            
        # Use a pre-created spreadsheet ID if none provided
        if not spreadsheet_id:
            print("⚠️  No spreadsheet ID provided. Will attempt to create a new one.")
            
        try:
            sheet_name = f"Persona_{username}_{datetime.now().strftime('%Y%m%d')}"
            values = [
                ["Category", "Detail", "Source"],
                ["BASIC INFORMATION", "", ""],
                ["Username", username, ""],
                ["Age", persona_data.get('age', 'N/A'), ""],
                ["Location", persona_data.get('location', 'N/A'), ""],
                ["Occupation", persona_data.get('occupation', 'N/A'), ""],
                ["Relationship Status", persona_data.get('marriage_status', 'N/A'), ""],
                ["PERSONALITY & ARCHETYPE", "", ""],
                ["Archetype", persona_data.get('archetype', 'N/A'), ""],
                ["Communication Style", persona_data.get('personality', 'N/A'), ""],
                ["MOTIVATIONS", "", ""],
            ]
            
            # Add motivations
            for i, (motivation, source) in enumerate(persona_data.get('motivations', [])):
                values.append([f"Motivation {i+1}", motivation, source])
                
            values.append(["GOALS & NEEDS", "", ""])
            for i, (goal, source) in enumerate(persona_data.get('goals', [])):
                values.append([f"Goal {i+1}", goal, source])
                
            values.append(["BEHAVIOR & HABITS", "", ""])
            for i, (behavior, source) in enumerate(persona_data.get('behaviors', [])):
                values.append([f"Behavior {i+1}", behavior, source])
                
            values.append(["FRUSTRATIONS", "", ""])
            for i, (frustration, source) in enumerate(persona_data.get('frustrations', [])):
                values.append([f"Frustration {i+1}", frustration, source])
                
            values.extend([
                ["ACTIVITY SUMMARY", "", ""],
                ["Activity Level", persona_data.get('activity_level', 'N/A'), ""],
                ["Total Comments", persona_data.get('total_comments', 0), ""],
                ["Total Posts", persona_data.get('total_posts', 0), ""],
            ])
            
            # Create a new spreadsheet if no ID provided
            try:
                # Try to create a new spreadsheet
                print("Attempting to create a new spreadsheet...")
                spreadsheet = {
                    'properties': {
                        'title': f'Reddit Persona - {username} - {datetime.now().strftime("%Y-%m-%d")}'
                    }
                }
                spreadsheet = self.google_sheets_service.spreadsheets().create(
                    body=spreadsheet,
                    fields='spreadsheetId,spreadsheetUrl'
                ).execute()
                spreadsheet_id = spreadsheet.get('spreadsheetId')
                print(f"✅ Created new spreadsheet with ID: {spreadsheet_id}")
            except Exception as e:
                print(f"❌ Error creating spreadsheet: {str(e)}")
                print("\nTroubleshooting steps:")
                print("1. Go to https://console.cloud.google.com/")
                print("2. Select your project")
                print("3. Go to 'APIs & Services' > 'Library'")
                print("4. Search for and enable 'Google Sheets API'")
                print("5. Go to 'IAM & Admin' > 'Service Accounts'")
                print("6. Find your service account and add 'Editor' role")
                print("7. Wait a few minutes for changes to take effect")
                return None
            
            # Update the sheet
            body = {'values': values}
            result = self.google_sheets_service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f"{sheet_name}!A1",
                valueInputOption='RAW',
                body=body
            ).execute()
            
            return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
            
        except Exception as e:
            print(f"Error exporting to Google Sheets: {e}")
            return None
            

        
    def extract_username(self, input_text):
        """Extract username from input text (can be URL or just username)"""
        if not input_text:
            return None
            
        # Remove any whitespace
        input_text = input_text.strip()
        
        # If it's just a username (with or without u/)
        if not any(c in input_text for c in ['/', '.', ':']):
            return input_text.lstrip('u/')
            
        # Handle different URL formats
        if 'reddit.com/user/' in input_text:
            return input_text.split('reddit.com/user/')[-1].split('/')[0].split('?')[0]
        elif 'redd.it/' in input_text:
            return input_text.split('redd.it/')[-1].split('/')[0]
        elif input_text.startswith(('http://', 'https://')):
            # Handle other URL formats by taking the last part
            return input_text.rstrip('/').split('/')[-1]
        else:
            # If it doesn't match any pattern, assume it's a username
            return input_text.lstrip('u/')
        
    def clean_text(self, text):
        """Clean and tokenize text for analysis"""
        try:
            if not text or not isinstance(text, str):
                return ""
                
            # Remove URLs, special characters, and numbers
            text = re.sub(r'http\S+|www\.\S+', '', text)
            text = re.sub(r'[^\w\s]', ' ', text.lower())
            
            # Tokenize and remove stopwords
            stop_words = set(stopwords.words('english'))
            words = word_tokenize(text)
            return ' '.join([word for word in words if word not in stop_words and len(word) > 2])
        except Exception as e:
            print(f"\nError cleaning text: {e}")
            return ""
    
    def analyze_sentiment(self, text):
        """Analyze sentiment of a text"""
        return self.sia.polarity_scores(text)
    
    def get_common_words(self, texts, n=10):
        """Get most common words from a list of texts"""
        words = []
        for text in texts:
            words.extend(text.split())
        return Counter(words).most_common(n)
    
    def get_user_comments(self, username, limit=100):
        """Fetch user's comments from Reddit"""
        try:
            user = self.reddit.redditor(username)
            comments = []
            for comment in user.comments.new(limit=limit):
                comments.append({
                    'body': comment.body,
                    'subreddit': comment.subreddit.display_name,
                    'score': comment.score,
                    'created_utc': comment.created_utc,
                    'url': f"https://reddit.com{comment.permalink}"
                })
            return comments
        except Exception as e:
            print(f"Error fetching comments: {e}")
            return []

    def get_user_posts(self, username, limit=50):
        """Fetch user's posts from Reddit"""
        try:
            user = self.reddit.redditor(username)
            posts = []
            for submission in user.submissions.new(limit=limit):
                posts.append({
                    'title': submission.title,
                    'selftext': submission.selftext,
                    'subreddit': submission.subreddit.display_name,
                    'score': submission.score,
                    'created_utc': submission.created_utc,
                    'url': f"https://reddit.com{submission.permalink}",
                    'is_self': submission.is_self,
                    'num_comments': submission.num_comments
                })
            return posts
        except Exception as e:
            print(f"Error fetching posts: {e}")
            return []

    def analyze_user(self, username_or_url, limit=100, export_to_sheets=False):
        """Analyze a Reddit user and generate a persona."""
        username = self.extract_username(username_or_url)
        if not username:
            print("Error: No username provided.")
            return None
            
        print(f"\nAnalyzing u/{username}...")
        
        try:
            # Get user comments and posts
            comments = self.get_user_comments(username, limit)
            posts = self.get_user_posts(username, limit // 2)  # Fewer posts than comments
            
            # Combine all text for analysis
            all_text = " ".join([comment['body'] for comment in comments] + 
                              [post.get('title', '') + " " + post.get('selftext', '') 
                               for post in posts if 'title' in post])
            
            # Analyze sentiment
            avg_sentiment = self.analyze_sentiment(all_text)
            
            # Generate and print persona
            print("\nGenerating persona summary...")
            try:
                persona = self.generate_persona_summary(username, comments, posts, avg_sentiment)
                
                if not persona:
                    print(f"❌ Failed to generate persona for u/{username}")
                    return None
                    
                print("✅ Persona generated successfully!")
                print("\n" + "="*50)
                print(f"📊 PERSONA SUMMARY: u/{username}".center(50))
                print("="*50)
            except Exception as e:
                print(f"Error generating persona: {str(e)}")
                import traceback
                traceback.print_exc()
                return None
                
            # Ensure we have all required fields with defaults
            default_dict = {
                'age': 'Not specified',
                'location': 'Not specified',
                'occupation': 'Not specified',
                'marriage_status': 'Not specified',
                'archetype': 'Not specified',
                'communication_style': 'Not specified',
                'motivations': [],
                'goals': [],
                'behaviors': [],
                'frustrations': [],
                'activity_level': 'Not active',
                'top_subreddits': []
            }
            
            # Create the persona data with defaults
            persona_data = {'username': username}
            for key, default in default_dict.items():
                persona_data[key] = persona.get(key, default)
                
            # Add comment and post counts
            persona_data.update({
                'total_comments': len(comments),
                'total_posts': len(posts)
            })
            
            # Export to Google Sheets if requested
            if export_to_sheets and hasattr(self, 'google_sheets_service') and self.google_sheets_service:
                self.export_to_google_sheets(username, persona_data)
                
            return persona_data
            
        except Exception as e:
            print(f"An error occurred while analyzing u/{username}: {e}")
            if hasattr(e, 'response') and hasattr(e.response, 'text'):
                print(f"Response: {e.response.text}")
            return None
            
    def extract_motivations(self, comments, posts):
        """Extract user motivations from comments and posts"""
        motivations = []
        for comment in comments:
            text = comment['body'].lower()
            if any(word in text for word in ['want to', 'hope to', 'aspire to', 'dream of', 'goal is']):
                motivations.append((
                    "Wants to " + text.split('want to')[-1][:100] + "...",
                    f"Comment in r/{comment['subreddit']}"
                ))
        return motivations or [("No explicit motivations mentioned recently", "N/A")]

    def extract_goals(self, comments, posts):
        """Extract user goals from comments and posts"""
        goals = []
        for post in posts:
            text = (post.get('title', '') + ' ' + post.get('selftext', '')).lower()
            if any(word in text for word in ['goal', 'objective', 'aim', 'target']):
                goals.append((
                    "Aims to " + text.split('goal')[-1][:100] + "...",
                    f"Post in r/{post['subreddit']}"
                ))
        return goals or [("Not explicitly mentioned in recent activity", "N/A")]

    def extract_behaviors(self, comments, posts):
        """Extract user behaviors from comments and posts"""
        behaviors = []
        for comment in comments:
            text = comment['body'].lower()
            if 'i always' in text or 'i usually' in text or 'i often' in text:
                behaviors.append((
                    "Habit: " + text[:150] + "...",
                    f"Comment in r/{comment['subreddit']}"
                ))
        return behaviors or [("Patterns not clearly identifiable from recent activity", "N/A")]

    def extract_frustrations(self, comments, posts):
        """Extract user frustrations from comments and posts"""
        frustrations = []
        for comment in comments:
            text = comment['body'].lower()
            if any(word in text for word in ['frustrat', 'annoy', 'bother', 'problem', 'issue']):
                frustrations.append((
                    "Frustrated by: " + text[:150] + "...",
                    f"Comment in r/{comment['subreddit']}"
                ))
        return frustrations or [("No explicit frustrations mentioned recently", "N/A")]

    def determine_archetype(self, comments, posts, personality_trait):
        """Determine user archetype based on activity"""
        if not comments and not posts:
            return "Observer"
        
        # Count different types of interactions
        question_count = sum(1 for c in comments if '?' in c['body'])
        answer_count = sum(1 for c in comments if '?' not in c['body'] and len(c['body'].split()) > 10)
        
        if question_count > answer_count * 2:
            return "The Inquirer"
        elif answer_count > question_count * 2:
            return "The Helper"
        elif len(set(c['subreddit'] for c in comments)) > 5:
            return "The Explorer"
        else:
            return "The Engaged Member"

    def extract_personal_info(self, text, comments, posts):
        """Extract personal information from user's activity"""
        info = {
            'age': 'Not specified',
            'occupation': 'Not specified',
            'marriage_status': 'Not specified',
            'location': 'Not specified'
        }
        
        # Combine all content for analysis
        all_content = [text] + [c['body'] for c in comments] + \
                     [p.get('title', '') + ' ' + p.get('selftext', '') for p in posts]
        
        # Look for age mentions with specific patterns
        age_patterns = [
            # Matches "I'm X years old" or "I am X years old"
            r'(?:i[\'\'\’]m|i am|age is|aged|turning)\s+(\d{1,2})\s*(?:years?\s*old|y\/o|yo|y\.o\.|\b)',
            # Matches "age X" or "aged X"
            r'(?:age|aged)\s+(\d{1,2})\s*(?:years?\s*old|y\/o|yo|y\.o\.|\b)',
            # Matches "X years old" or "X yo"
            r'(\d{1,2})\s*(?:years?\s*old|y\/o|yo|y\.o\.)(?:\b|\W|$)',
            # Matches "born in 19XX" or "born in 20XX"
            r'born (?:in|on|\w+)?\s*(?:the year )?(?:of )?(19\d{2}|20[01]\d)(?:\D|$)',
            # Matches "turned X last year" or "when I was X"
            r'(?:turned|when i was|since i was|since age|age)\s+(\d{1,2})\b'
        ]
        
        current_year = datetime.now().year
        
        for content in all_content:
            # Clean the content (remove markdown, URLs, etc.)
            clean_content = re.sub(r'\[.*?\]\(.*?\)', '', content)  # Remove markdown links
            clean_content = re.sub(r'http\S+', '', clean_content)  # Remove URLs
            
            # Check for age patterns
            for pattern in age_patterns:
                match = re.search(pattern, clean_content, re.IGNORECASE)
                if match:
                    age = None
                    # Handle different group patterns
                    if len(match.groups()) > 1:
                        age = next((g for g in match.groups() if g and g.isdigit() and 13 <= int(g) <= 100), None)
                    elif match.group(1):
                        if match.group(1).isdigit() and 13 <= int(match.group(1)) <= 100:
                            age = match.group(1)
                    
                    # Special handling for birth years
                    if not age and len(match.groups()) > 0 and match.group(1):
                        year = match.group(1)
                        if year.isdigit() and len(year) == 4 and 1900 <= int(year) <= current_year:
                            age = str(current_year - int(year))
                    
                    if age and 13 <= int(age) <= 100:  # Reasonable age range
                        info['age'] = f"{age} years old"
                        # Try to find a more specific age if available
                        if 'years old' not in clean_content.lower() and int(age) < 30:
                            # If age is approximate, try to find more specific mentions
                            specific_age = re.search(rf'{age}\s*(?:years?\s*old|y\/o|yo|y\.o\.)', clean_content, re.IGNORECASE)
                            if specific_age:
                                info['age'] = f"{age} years old"
                        break
            
            if info['age'] != 'Not specified':
                break
        
        # Look for location mentions with more specific patterns
        location_indicators = [
            'i live in', 'i\'m from', 'i am from', 'based in', 'located in', 
            'hometown', 'currently in', 'living in', 'reside in', 'based out of'
        ]
        
        # List of common locations to match against
        common_locations = [
            # US States
            'Alabama', 'Alaska', 'Arizona', 'Arkansas', 'California', 'Colorado', 'Connecticut', 'Delaware',
            'Florida', 'Georgia', 'Hawaii', 'Idaho', 'Illinois', 'Indiana', 'Iowa', 'Kansas', 'Kentucky',
            'Louisiana', 'Maine', 'Maryland', 'Massachusetts', 'Michigan', 'Minnesota', 'Mississippi', 'Missouri',
            'Montana', 'Nebraska', 'Nevada', 'New Hampshire', 'New Jersey', 'New Mexico', 'New York',
            'North Carolina', 'North Dakota', 'Ohio', 'Oklahoma', 'Oregon', 'Pennsylvania', 'Rhode Island',
            'South Carolina', 'South Dakota', 'Tennessee', 'Texas', 'Utah', 'Vermont', 'Virginia', 'Washington',
            'West Virginia', 'Wisconsin', 'Wyoming',
            # Countries
            'United States', 'Canada', 'United Kingdom', 'Australia', 'Germany', 'France', 'Japan', 'China',
            'India', 'Brazil', 'Mexico', 'Italy', 'Spain', 'Russia', 'South Korea'
        ]
        
        # Common location patterns (cities, states, countries)
        location_patterns = [
            # Matches "I live in [Location]" or "Based in [Location]" etc.
            r'(?:{})\s+([A-Z][A-Za-z\s]{{2,}}(?:,\s*[A-Z][A-Za-z\s]+)*)'.format('|'.join(location_indicators)),
            # Matches "in [City, State]" or "in [City, Country]"
            r'in\s+([A-Z][A-Za-z\s]+(?:,\s*(?:[A-Z][a-z]+\s*)+)?)',
            # Matches "from [Location]"
            r'from\s+([A-Z][A-Za-z\s]+(?:,\s*[A-Za-z\s]+)*)',
            # Matches common locations as standalone words
            r'\b({})\b'.format('|'.join(common_locations))
        ]
        
        # Check all content for location mentions
        for content in all_content:
            # Skip very short content to avoid false positives
            if len(content) < 30:  # Increased minimum length
                continue
                
            # Clean the content (remove markdown, URLs, etc.)
            clean_content = re.sub(r'\[.*?\]\(.*?\)', '', content)  # Remove markdown links
            clean_content = re.sub(r'http\S+', '', clean_content)  # Remove URLs
            
            for pattern in location_patterns:
                matches = re.finditer(pattern, clean_content, re.IGNORECASE)
                for match in matches:
                    location = match.group(1).strip() if match.lastindex else match.group(0).strip()
                    
                    # Skip common false positives
                    if len(location) < 3:
                        continue
                        
                    # Additional validation
                    if any(word in location.lower() for word in [
                        'reddit', 'internet', 'online', 'web', 'here', 'today', 
                        'yesterday', 'tomorrow', 'this', 'that', 'there'
                    ]):
                        continue
                        
                    # If it's a common location or matches a specific pattern
                    if (any(loc.lower() == location.lower() for loc in common_locations) or
                        re.match(r'^[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*(?:,\s*[A-Z][a-z]+)?$', location)):
                        info['location'] = location
                        break
                        
                if info['location'] != 'Not specified':
                    break
                    
            if info['location'] != 'Not specified':
                break
        
        # Enhanced occupation extraction
        occupation_indicators = [
            # Work-related phrases
            r'(?:i(?:\'?m| am|\'ve been)?\s+(?:working\s+)?(?:as|at|in))\s+([a-z\s-]+(?:\s+at\s+[a-z\s-]+)?)',
            r'(?:my\s+(?:current\s+)?(?:job|profession|occupation|role|position|title)(?:\s+is|\:))\s+([a-z\s-]+)',
            r'(?:i(?:\'?m| am| work)?\s+(?:a|an|the)?\s*)([a-z\s-]+(?:\s+by\s+[a-z\s-]+)?)(?:\s+by\s+profession|\s+by\s+trade|\s+here|\s+myself)',
            # Education-related
            r'(?:i(?:\'?m| am| study|\'m studying| study|major(?:ing)? in|majored in|pursuing(?: a degree in)?))\s+([a-z\s-]+(?:\s+at\s+[a-z\s-]+)?)',
            r'(?:i(?:\'?m| am| was)?\s+an?\s+)([a-z\s-]+(?:\s+student\b|\s+at\b|\s+in\b|\s+studying\b))',
            # Industry-specific
            r'(?:i(?:\'?m| am| work)?\s+in\s+(?:the\s+)?)([a-z\s-]+(?:\s+industry|\s+field|\s+sector|\s+area))',
            # Self-employed/Entrepreneur
            r'(?:i(?:\'?m| am| run)?\s+(?:a|an|my|the)?\s*)([a-z\s-]+(?:\s+business|\s+company|\s+startup|\s+venture|\s+shop))',
            r'(?:i(?:\'?m| am| own| operate)?\s+(?:a|an|my|the)?\s*)([a-z\s-]+(?:\s+store|\s+shop|\s+service))',
            # Freelance/Contract work
            r'(?:i(?:\'?m| am)?\s+(?:a|an)?\s*)([a-z\s-]+(?:\s+freelance\w*|\s+contractor|\s+consultant))',
            # Retired/Unemployed
            r'(i(?:\'?m| am)\s+(?:a\s+)?(?:retired|unemployed|between jobs|looking for work|job hunting|seeking employment))',
            # Generic fallback
            r'(?:i(?:\'?m| am| work)?\s+(?:as\s+)?)([a-z\s-]+(?:\s+at\s+[a-z\s-]+)?)'
        ]
        
        # Common job titles and fields to validate against
        common_occupations = [
            # Professional/White-collar
            'engineer', 'developer', 'programmer', 'designer', 'analyst', 'manager', 'director', 'executive',
            'consultant', 'architect', 'scientist', 'researcher', 'professor', 'teacher', 'instructor',
            'lawyer', 'attorney', 'doctor', 'physician', 'nurse', 'therapist', 'counselor', 'accountant',
            'marketer', 'specialist', 'strategist', 'planner', 'coordinator', 'administrator',
            # Trades/Blue-collar
            'technician', 'mechanic', 'electrician', 'plumber', 'carpenter', 'contractor', 'builder',
            'chef', 'cook', 'baker', 'bartender', 'server', 'waiter', 'waitress', 'barista',
            'driver', 'operator', 'laborer', 'factory worker', 'warehouse worker', 'delivery driver',
            # Creative
            'artist', 'writer', 'author', 'musician', 'actor', 'actress', 'performer', 'photographer',
            'filmmaker', 'producer', 'editor', 'journalist', 'reporter', 'blogger', 'influencer',
            # Service
            'sales', 'retail', 'cashier', 'customer service', 'receptionist', 'assistant', 'secretary',
            'hairdresser', 'stylist', 'esthetician', 'masseuse', 'trainer', 'coach', 'instructor',
            # Other
            'student', 'researcher', 'scientist', 'analyst', 'entrepreneur', 'business owner', 'freelancer',
            'consultant', 'contractor', 'self-employed', 'retired', 'unemployed', 'homemaker', 'parent'
        ]
        
        # Check all content for occupation mentions
        for content in all_content:
            # Skip very short content to avoid false positives
            if len(content) < 30:
                continue
                
            # Clean the content (remove markdown, URLs, etc.)
            clean_content = re.sub(r'\[.*?\]\(.*?\)', '', content)  # Remove markdown links
            clean_content = re.sub(r'http\S+', '', clean_content)  # Remove URLs
            clean_content = re.sub(r'[^\w\s]', ' ', clean_content)  # Remove special chars
            clean_content = re.sub(r'\s+', ' ', clean_content).strip().lower()  # Normalize whitespace
            
            # Check for occupation patterns
            for pattern in occupation_indicators:
                matches = re.finditer(pattern, clean_content, re.IGNORECASE)
                for match in matches:
                    occupation = match.group(1).strip() if match.lastindex else match.group(0).strip()
                    
                    # Clean up the extracted occupation
                    occupation = re.sub(r'^[^a-z]+', '', occupation)  # Remove leading non-letters
                    occupation = re.sub(r'\b(?:a|an|the|my|at|in|for|with|and|or|but)\b', '', occupation)  # Remove common words
                    occupation = re.sub(r'\s+', ' ', occupation).strip()  # Normalize whitespace
                    
                    # Skip if too short or contains invalid characters
                    if len(occupation) < 3 or len(occupation.split()) > 5:
                        continue
                        
                    # Check if any common occupation is mentioned
                    if any(occ in occupation for occ in common_occupations):
                        info['occupation'] = occupation.title()
                        break
                        
                    # Additional validation for job titles
                    if (re.match(r'^[a-z]+(?:\s+[a-z]+){0,3}$', occupation) and 
                        not any(word in occupation for word in ['i', 'me', 'my', 'you', 'your', 'we', 'us', 'our'])):
                        info['occupation'] = occupation.title()
                        break
                
                if info['occupation'] != 'Not specified':
                    break
            
            if info['occupation'] != 'Not specified':
                break
                
        # If no occupation found, try to infer from subreddits
        if info['occupation'] == 'Not specified' and comments:
            # Get top subreddits
            subreddits = [c['subreddit'].lower() for c in comments]
            subreddit_counter = Counter(subreddits)
            
            # Check if user is active in career/professional subreddits
            professional_subs = {
                'programming', 'webdev', 'learnprogramming', 'cscareerquestions', 'datascience',
                'engineering', 'askengineers', 'medicine', 'nursing', 'law', 'lawyers', 'teachers',
                'marketing', 'sales', 'entrepreneur', 'startups', 'smallbusiness', 'freelance',
                'graphic_design', 'photography', 'filmmakers', 'writing', 'art', 'music', 'chefs',
                'talesfromyourserver', 'talesfromretail', 'talesfromtechsupport'
            }
            
            # Find matching professional subreddits
            user_professional_subs = [sub for sub in subreddit_counter if sub in professional_subs]
            if user_professional_subs:
                # Get the most active professional subreddit
                most_active = max(user_professional_subs, key=lambda x: subreddit_counter[x])
                
                # Map subreddits to occupations
                sub_to_occupation = {
                    'programming': 'Software Developer',
                    'webdev': 'Web Developer',
                    'learnprogramming': 'Aspiring Programmer',
                    'cscareerquestions': 'Tech Professional',
                    'datascience': 'Data Scientist',
                    'engineering': 'Engineer',
                    'askengineers': 'Engineer',
                    'medicine': 'Medical Professional',
                    'nursing': 'Nurse',
                    'law': 'Legal Professional',
                    'lawyers': 'Lawyer',
                    'teachers': 'Teacher',
                    'marketing': 'Marketing Professional',
                    'sales': 'Sales Professional',
                    'entrepreneur': 'Entrepreneur',
                    'startups': 'Startup Founder',
                    'smallbusiness': 'Small Business Owner',
                    'freelance': 'Freelancer',
                    'graphic_design': 'Graphic Designer',
                    'photography': 'Photographer',
                    'filmmakers': 'Filmmaker',
                    'writing': 'Writer',
                    'art': 'Artist',
                    'music': 'Musician',
                    'chefs': 'Chef',
                    'talesfromyourserver': 'Restaurant Server',
                    'talesfromretail': 'Retail Worker',
                    'talesfromtechsupport': 'IT Support'
                }
                
                info['occupation'] = sub_to_occupation.get(most_active, 'Not specified')
        
        # If still no occupation found, check for student status
        if info['occupation'] == 'Not specified':
            student_indicators = [
                'college student', 'university student', 'grad student', 'graduate student',
                'high school student', 'student at', 'studying at', 'pursuing', 'majoring in'
            ]
            
            for content in all_content:
                if any(indicator in content.lower() for indicator in student_indicators):
                    info['occupation'] = 'Student'
                    break
        
        # Enhanced relationship status detection
        relationship_patterns = [
            # Married/Partnered
            (r'(?:i(?:\'?m| am|\'ve been)?\s+(?:happily\s+)?(?:married|engaged|betrothed|wed(?:ded)?)(?:\s+to\s+\w+)?(?:\s+for\s+\w+)?\b)', 'married'),
            (r'(?:my\s+(?:wife|husband|spouse|fianc[ée]e?)(?:\s+and\s+i)?\b)', 'married'),
            (r'(?:we(?:\'?re| are)?\s+(?:married|engaged|together)(?:\s+for\s+\w+)?\b)', 'married'),
            
            # In a relationship
            (r'(?:i(?:\'?m| am|\'ve been)?\s+(?:in\s+a\s+relationship|dating|going\s+out|seeing\s+someone)(?:\s+with\s+\w+)?(?:\s+for\s+\w+)?\b)', 'in a relationship'),
            (r'(?:my\s+(?:girlfriend|boyfriend|partner|s.o.|significant\s+other|better\s+half)(?:\s+and\s+i)?\b)', 'in a relationship'),
            (r'(?:we(?:\'?re| are)?\s+together(?:\s+for\s+\w+)?\b)', 'in a relationship'),
            
            # Single/Dating
            (r'(?:i(?:\'?m| am)?\s+(?:single|unattached|not\s+seeing\s+anyone|not\s+dating(?:\s+anyone)?|not\s+in\s+a\s+relationship)\b)', 'single'),
            (r'(?:i(?:\'?m| am)?\s+(?:dating\s+around|playing\s+the\s+field|happily\s+single))', 'single'),
            
            # Complicated/Other
            (r'(?:it\'s\s+complicated|complicated\s+relationship|on\s+and\s+off)', 'it\'s complicated'),
            (r'(?:in\s+an?\s+open\s+relationship|open\s+marriage|ethically\s+non\-?monogamous)', 'in an open relationship'),
            (r'(?:divorc(?:ed|ing)|separat(?:ed|ing)|split\s+up|broke\s+up)', 'divorced/separated'),
            (r'(?:widow(?:ed)?|lost\s+my\s+(?:wife|husband|partner))', 'widowed')
        ]
        
        # Additional context patterns that might indicate relationship status
        context_patterns = [
            (r'(?:my\s+(?:wife|husband|spouse|fianc[ée]e?|girlfriend|boyfriend|partner|s.o.|significant\s+other))', 'in a relationship'),
            (r'(?:our\s+(?:anniversary|wedding|marriage|relationship))', 'married'),
            (r'(?:we\'ve\s+been\s+together\s+for)', 'in a relationship'),
            (r'(?:my\s+ex(?:\-\w+)?\b)', 'single')
        ]
        
        # Check all content for relationship status mentions
        for content in all_content:
            # Clean the content
            clean_content = re.sub(r'\[.*?\]\(.*?\)', '', content)  # Remove markdown links
            clean_content = re.sub(r'http\S+', '', clean_content)  # Remove URLs
            clean_content = clean_content.lower()
            
            # Check direct relationship patterns
            for pattern, status in relationship_patterns:
                if re.search(pattern, clean_content, re.IGNORECASE):
                    info['marriage_status'] = status
                    break
            
            # If no direct match, check contextual patterns
            if info['marriage_status'] == 'Not specified':
                for pattern, status in context_patterns:
                    if re.search(pattern, clean_content, re.IGNORECASE):
                        info['marriage_status'] = status
                        break
            
            if info['marriage_status'] != 'Not specified':
                break
                
        # If still not specified, check for family-related subreddits
        if info['marriage_status'] == 'Not specified' and comments:
            family_subs = {
                'marriage', 'relationships', 'relationship_advice', 'dating_advice',
                'weddingplanning', 'weddings', 'divorce', 'singleparents', 'dating'
            }
            
            user_subs = {c['subreddit'].lower() for c in comments}
            if user_subs & family_subs:
                if 'divorce' in user_subs:
                    info['marriage_status'] = 'divorced/separated'
                elif 'singleparents' in user_subs:
                    info['marriage_status'] = 'single parent'
                elif any(sub in user_subs for sub in ['marriage', 'weddingplanning', 'weddings']):
                    info['marriage_status'] = 'married'
                elif any(sub in user_subs for sub in ['relationships', 'dating_advice', 'dating']):
                    info['marriage_status'] = 'in a relationship'
        
        return info

    def safe_print(self, text):
        """Safely print text with emoji and unicode support"""
        try:
            print(text)
        except UnicodeEncodeError:
            # Fallback for environments with limited encoding support
            cleaned = text.encode('ascii', 'replace').decode('ascii')
            print(cleaned)

    def generate_persona_summary(self, username, comments, posts, avg_sentiment, export_to_sheets=False):
        """Generate a detailed user persona with citations and optional Google Sheets export"""
        # Combine all text for analysis
        all_text = " ".join([comment['body'] for comment in comments] + 
                          [post.get('title', '') + " " + post.get('selftext', '') 
                           for post in posts if 'title' in post])
        
        # Extract personal information
        personal_info = self.extract_personal_info(all_text.lower(), comments, posts)
        
        # Get persona elements
        motivations = self.extract_motivations(comments, posts)
        goals = self.extract_goals(comments, posts)
        behaviors = self.extract_behaviors(comments, posts)
        frustrations = self.extract_frustrations(comments, posts)
        
        # Determine personality traits using the compound score from sentiment analysis
        compound_score = avg_sentiment.get('compound', 0)  # Default to neutral if compound score not found
        personality = "Friendly and engaging" if compound_score > 0.1 else \
                     "Neutral or reserved" if -0.1 <= compound_score <= 0.1 else \
                     "Direct or critical"
        
        # Determine archetype
        archetype = self.determine_archetype(comments, posts, personality)
        
        # Activity metrics
        total_comments = len(comments)
        total_posts = len(posts)
        activity_level = "Highly active" if total_comments + total_posts > 200 else \
                        "Active" if total_comments + total_posts > 50 else "Casual"
        
        # Prepare persona data for potential export
        persona_data = {
            'username': username,
            'age': personal_info['age'],
            'location': personal_info['location'],
            'occupation': personal_info['occupation'],
            'marriage_status': personal_info['marriage_status'],
            'archetype': archetype,
            'personality': personality,
            'motivations': motivations[:3],
            'goals': goals[:3],
            'behaviors': behaviors[:3],
            'frustrations': frustrations[:3],
            'activity_level': activity_level,
            'total_comments': total_comments,
            'total_posts': total_posts
        }
        
        # Print the detailed persona with enhanced formatting
        self.safe_print("\n" + "✨" + "="*78 + "✨")
        self.safe_print(f"🔍  REDDIT PERSONA ANALYSIS: u/{username}".center(80))
        self.safe_print("✨" + "="*78 + "✨\n")
        
        # Basic Information with emojis and better formatting
        self.safe_print("\n\U0001f4cb " + "BASIC INFORMATION".ljust(75, '\u2500'))
        self.safe_print(f"   👤 Username: u/{username}")
        self.safe_print(f"   🎂 Age: {personal_info['age']}")
        self.safe_print(f"   📍 Location: {personal_info['location']}")
        self.safe_print(f"   💼 Occupation: {personal_info['occupation']}")
        self.safe_print(f"   💑 Relationship Status: {personal_info['marriage_status']}")
        
        # Personality & Archetype with visual indicators
        self.safe_print("\n\U0001f9e0 " + "PERSONALITY & ARCHETYPE".ljust(75, '\u2500'))
        self.safe_print(f"   🧩 Archetype: {archetype}")
        self.safe_print(f"   🧠 Personality: {personality}")
        sentiment_emoji = '😊' if compound_score > 0.1 else '😐' if compound_score > -0.1 else '😟'
        self.safe_print(f"   {sentiment_emoji} Overall Sentiment: {abs(compound_score)*100:.1f}% {'positive' if compound_score > 0 else 'negative' if compound_score < 0 else 'neutral'}")
        
        # Motivations with emojis and better formatting
        self.safe_print("\n\U0001f4a1 " + "MOTIVATIONS".ljust(75, '\u2500'))
        for i, (motivation, source) in enumerate(motivations[:3], 1):
            self.safe_print(f"   {i}. {motivation}")
            self.safe_print(f"      📌 Source: {source}")
        
        # Goals & Needs with emojis and better formatting
        self.safe_print("\n\U0001f3af " + "GOALS & NEEDS".ljust(75, '\u2500'))
        for i, (goal, source) in enumerate(goals[:3], 1):
            self.safe_print(f"   {i}. {goal}")
            self.safe_print(f"      📌 Source: {source}")
        
        # Behavior & Habits with emojis and better formatting
        self.safe_print("\n\U0001f4dd " + "BEHAVIORS & HABITS".ljust(75, '\u2500'))
        for i, (behavior, source) in enumerate(behaviors[:3], 1):
            self.safe_print(f"   {i}. {behavior}")
            self.safe_print(f"      📌 Source: {source}")
        
        # Frustrations with emojis and better formatting
        self.safe_print("\n\U0001f621 " + "FRUSTRATIONS".ljust(75, '\u2500'))
        for i, (frustration, source) in enumerate(frustrations[:3], 1):
            self.safe_print(f"   {i}. {frustration}")
            self.safe_print(f"      📌 Source: {source}")
        
        # Activity Summary with emojis and better formatting
        self.safe_print("\n\U0001f4ca " + "ACTIVITY SUMMARY".ljust(75, '\u2500'))
        self.safe_print(f"   📊 Activity Level: {activity_level}")
        self.safe_print(f"   💬 Total Comments: {total_comments:,}")
        self.safe_print(f"   📝 Total Posts: {total_posts:,}")
        self.safe_print(f"   📅 Total Activity: {total_comments + total_posts:,} interactions")
        
        # Most active subreddits
        if comments or posts:
            all_subs = [comment['subreddit'] for comment in comments] + \
                      [post['subreddit'] for post in posts]
            if all_subs:
                common_subs = Counter(all_subs).most_common(5)
                self.safe_print("\n🏆 " + "MOST ACTIVE COMMUNITIES".ljust(75, '─'))
                for i, (sub, count) in enumerate(common_subs, 1):
                    bar = '█' * min(10, int((count / common_subs[0][1]) * 10))
                    self.safe_print(f"   {i}. r/{sub:<20} {bar} {count:,} interactions")
                    
                    # Add to persona data for export
                    persona_data['top_subreddits'] = [{'subreddit': sub, 'count': count} for sub, count in common_subs]
        
        # Export to Google Sheets if requested
        if export_to_sheets and hasattr(self, 'google_sheets_service') and self.google_sheets_service:
            sheet_url = self.export_to_google_sheets(username, persona_data)
            if sheet_url:
                self.safe_print("\n" + "="*80)
                
        return persona_data

def export_to_excel(username, persona_data):
    """Export persona data to an Excel file with proper formatting"""
    import pandas as pd
    from datetime import datetime
    import os
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    try:
        # Create output directory if it doesn't exist
        os.makedirs('exports', exist_ok=True)
        
        # Create a filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"reddit_persona_{username}_{timestamp}.xlsx"
        filepath = os.path.abspath(os.path.join('exports', filename))
        
        # Create a Pandas Excel writer using openpyxl
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Create a workbook and add a worksheet
            workbook = writer.book
            worksheet = workbook.create_sheet("Persona Analysis")
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            section_font = Font(bold=True, size=12, color="1F4E78")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Start row counter
            row = 1
            
            # Add title
            title = f"Reddit Persona Analysis - u/{username}"
            worksheet.append([title])
            worksheet.merge_cells(f'A{row}:B{row}')
            worksheet[f'A{row}'].font = Font(size=14, bold=True, color="1F4E78")
            worksheet[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2
            
            # Function to add a section
            def add_section(title):
                nonlocal row
                worksheet.append([title])
                worksheet.merge_cells(f'A{row}:B{row}')
                worksheet[f'A{row}'].font = section_font
                worksheet[f'A{row}'].fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                row += 1
            
            # Function to add key-value pairs
            def add_kv(key, value):
                nonlocal row
                worksheet.append([key, value])
                worksheet[f'A{row}'].font = Font(bold=True)
                worksheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
            
            # Add basic information
            add_section("🔹 BASIC INFORMATION")
            add_kv("Username:", f"u/{username}")
            add_kv("Age:", str(persona_data.get('age', 'N/A')))
            add_kv("Location:", str(persona_data.get('location', 'N/A')))
            add_kv("Occupation:", str(persona_data.get('occupation', 'N/A')))
            add_kv("Relationship Status:", str(persona_data.get('marriage_status', 'N/A')))
            row += 1  # Add empty row
            
            # Add personality section
            add_section("🧠 PERSONALITY & ARCHETYPE")
            add_kv("Archetype:", str(persona_data.get('archetype', 'N/A')))
            add_kv("Personality:", str(persona_data.get('personality', 'N/A')))
            row += 1
            
            # Add motivations
            motivations = persona_data.get('motivations', [])
            if motivations:
                add_section("💡 MOTIVATIONS")
                for i, (motivation, _) in enumerate(motivations, 1):
                    add_kv(f"{i}.", str(motivation))
                row += 1
            
            # Add goals
            goals = persona_data.get('goals', [])
            if goals:
                add_section("🎯 GOALS & NEEDS")
                for i, (goal, _) in enumerate(goals, 1):
                    add_kv(f"{i}.", str(goal))
                row += 1
            
            # Add behaviors
            behaviors = persona_data.get('behaviors', [])
            if behaviors:
                add_section("📝 BEHAVIORS & HABITS")
                for i, (behavior, _) in enumerate(behaviors, 1):
                    add_kv(f"{i}.", str(behavior))
                row += 1
            
            # Add frustrations
            frustrations = persona_data.get('frustrations', [])
            if frustrations:
                add_section("😡 FRUSTRATIONS")
                for i, (frustration, _) in enumerate(frustrations, 1):
                    add_kv(f"{i}.", str(frustration))
                row += 1
            
            # Add activity summary
            add_section("📊 ACTIVITY SUMMARY")
            add_kv("Activity Level:", str(persona_data.get('activity_level', 'N/A')))
            add_kv("Total Comments:", str(persona_data.get('total_comments', 0)))
            add_kv("Total Posts:", str(persona_data.get('total_posts', 0)))
            
            # Add top subreddits if available
            if 'top_subreddits' in persona_data and persona_data['top_subreddits']:
                add_section("🏆 TOP SUBREDDITS")
                for i, sub in enumerate(persona_data['top_subreddits'], 1):
                    if isinstance(sub, dict):
                        sub_name = sub.get('subreddit', 'N/A')
                        count = sub.get('count', 0)
                        add_kv(f"{i}. r/{sub_name}", f"{count} interactions")
                    else:
                        add_kv(f"{i}.", f"r/{sub}")
            
            # Adjust column widths
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
            
            # Apply borders to all cells
            for row_cells in worksheet.iter_rows():
                for cell in row_cells:
                    cell.border = border
            
            # Remove the default sheet created by pandas
            if 'Sheet' in workbook.sheetnames:
                std = workbook['Sheet']
                workbook.remove(std)
            
            # Save the workbook
            workbook.save(filepath)
        
        # Verify the file was created
        if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
            print(f"\n✅ Excel file created successfully: {filepath}")
            try:
                # Try to open with default application
                if os.name == 'nt':  # Windows
                    os.startfile(filepath)
                elif os.name == 'posix':  # macOS and Linux
                    if sys.platform == 'darwin':
                        os.system(f'open "{filepath}"')
                    else:
                        os.system(f'xdg-open "{filepath}"')
            except Exception as e:
                print(f"Note: Could not open file automatically: {e}")
                print("Please open the file manually from the exports folder.")
            return filepath
        else:
            raise Exception("Failed to create Excel file: File is empty or not created")
            
    except Exception as e:
        # Clean up if file creation failed
        if 'filepath' in locals() and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except:
                pass
        print(f"Error exporting to text file: {e}")
        raise  # Re-raise the exception to be handled by the caller

def main():
    import argparse
    from datetime import datetime
    
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Analyze Reddit user personas')
    parser.add_argument('username', nargs='?', help='Reddit username or profile URL')
    parser.add_argument('--export', action='store_true', help='Export to Google Sheets')
    parser.add_argument('--spreadsheet-id', help='Google Sheets spreadsheet ID (optional)')
    parser.add_argument('--excel', action='store_true', help='Export to Excel file (default if no export specified)')
    args = parser.parse_args()
    
    # Load environment variables
    load_dotenv()
    
    # Check for Google Sheets credentials
    google_creds_path = os.getenv('GOOGLE_CREDENTIALS_PATH', 'credentials.json')
    
    # Initialize the analyzer with Reddit API credentials
    analyzer = RedditPersonaAnalyzer(
        client_id=os.getenv('REDDIT_CLIENT_ID'),
        client_secret=os.getenv('REDDIT_CLIENT_SECRET'),
        user_agent='RedditPersonaAnalyzer/1.0',
        google_creds_path=google_creds_path
    )
    
    print("Reddit Persona Analyzer")
    print("======================\n")
    
    # Check if Google Sheets export is available
    export_to_sheets = False
    if hasattr(analyzer, 'google_sheets_service') and analyzer.google_sheets_service:
        export_to_sheets = True
        print("✅ Google Sheets export is available (credentials found)")
    else:
        print("ℹ️  Google Sheets export is not available. To enable:")
        print("   1. Create a Google Cloud Project and enable Google Sheets API")
        print("   2. Create service account credentials")
        print("   3. Save the JSON file as 'credentials.json' in this directory")
        print("   4. Or set GOOGLE_CREDENTIALS_PATH environment variable to the credentials file\n")
    
    try:
        while True:
            # Get input if not provided as argument
            if args.username:
                user_input = args.username
                args.username = None  # Clear for subsequent iterations
            else:
                print("\n" + "="*50)
                user_input = input("\nEnter Reddit username or profile URL (or 'q' to quit): ").strip()
            
            if not user_input or user_input.lower() == 'q':
                print("\nThank you for using Reddit Persona Analyzer!")
                break
            
            print(f"\nAnalyzing: {user_input}")
            
            # Analyze the user and get persona data
            persona_data = analyzer.analyze_user(user_input, export_to_sheets=False)
            
            if not persona_data:
                print("No data to export.")
                continue
                
            # Handle exports if requested
            if args.excel or (export_to_sheets and not args.excel) or not (args.excel or export_to_sheets):
                try:
                    excel_file = export_to_excel(user_input, persona_data)
                    print(f"\n Data exported to Excel: {os.path.abspath(excel_file)}")
                except Exception as e:
                    print(f"\n Error exporting to Excel: {str(e)}")
                    # Fallback to basic text export if Excel fails
                    try:
                        os.makedirs('exports', exist_ok=True)
                        filename = os.path.join('exports', f"reddit_persona_{user_input}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
                        with open(filename, 'w', encoding='utf-8') as f:
                            json.dump(persona_data, f, indent=2, ensure_ascii=False)
                        print(f" Data exported to JSON: {os.path.abspath(filename)}")
                    except Exception as e2:
                        print(f" Failed to export persona: {str(e2)}")
            
            if args.export or args.spreadsheet_id:
                print("\n" + "="*50)
                print("EXPORT OPTIONS")
                print("="*50)
                
                if args.spreadsheet_id or (hasattr(analyzer, 'google_sheets_service') and analyzer.google_sheets_service):
                    print("\nExporting to Google Sheets...")
                    if args.spreadsheet_id:
                        sheet_url = analyzer.export_to_google_sheets(user_input, persona_data, args.spreadsheet_id)
                    else:
                        sheet_url = analyzer.export_to_google_sheets(user_input, persona_data)
                    
                    if sheet_url:
                        print(f"✅ Data exported to Google Sheets: {sheet_url}")
                    else:
                        print("❌ Failed to export to Google Sheets. Falling back to CSV export.")
                        export_to_csv(user_input, persona_data)
                else:
                    print("Google Sheets export not available. Falling back to CSV export.")
                    export_to_csv(user_input, persona_data)
            
            # Always offer CSV export if not already done
            elif args.csv or (not args.export and not args.spreadsheet_id):
                print("\n" + "="*50)
                print("EXPORT OPTIONS")
                print("="*50)
                export_to_csv(user_input, persona_data)
            
            # Exit if username was provided as command line argument
            if len(sys.argv) > 1 and not any(opt in sys.argv[1:] for opt in ['--export', '--spreadsheet-id', '--csv']):
                break
                
    except KeyboardInterrupt:
        print("\nOperation cancelled by user.")
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        if hasattr(e, 'response') and hasattr(e.response, 'text'):
            print(f"Response: {e.response.text}")
        print("Please check the input and try again.")


if __name__ == "__main__":
    main()
